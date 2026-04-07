"""Формирование ДДС-отчёта в Excel."""
from datetime import datetime
from typing import List, Dict, Any, Tuple
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_dds_report(transactions: List[Dict[str, Any]], bank_name: str) -> Tuple[str, str]:
    wb = openpyxl.Workbook()
    ws_dds = wb.active
    ws_dds.title = "ДДС"
    build_dds_sheet(ws_dds, transactions)
    ws_txn = wb.create_sheet("Транзакции")
    build_transactions_sheet(ws_txn, transactions)
    report_path = f"/tmp/dds_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(report_path)
    return report_path, build_summary_text(transactions, bank_name)

def styled_cell(ws, row, col, value=None, fill=None, font=None, align="left", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill: c.fill = fill
    if font: c.font = font
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    if num_fmt: c.number_format = num_fmt
    return c

def build_dds_sheet(ws, transactions):
    monthly = defaultdict(lambda: defaultdict(float))
    for t in transactions:
        month = t["date"].strftime("%Y-%m")
        cat = t.get("category") or ("Прочие поступления" if t["amount"] > 0 else "Прочие расходы")
        monthly[month][cat] += t["amount"]
    months = sorted(monthly.keys())
    income_cats = sorted({c for md in monthly.values() for c, a in md.items() if a > 0})
    expense_cats = sorted({c for md in monthly.values() for c, a in md.items() if a < 0})

    H_FILL = PatternFill("solid", fgColor="1F4E79")
    IN_FILL = PatternFill("solid", fgColor="E2EFDA")
    EX_FILL = PatternFill("solid", fgColor="FCE4D6")
    TOT_FILL = PatternFill("solid", fgColor="D9E1F2")
    NDP_FILL = PatternFill("solid", fgColor="BDD7EE")
    H_FONT = Font(bold=True, color="FFFFFF")
    BOLD = Font(bold=True)
    N_COLS = len(months) + 2

    ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
    ws["A1"].value = "Отчёт о движении денежных средств (ДДС)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    row = 3
    styled_cell(ws, row, 1, "Статья", H_FILL, H_FONT, "center")
    for i, m in enumerate(months):
        styled_cell(ws, row, i+2, datetime.strptime(m, "%Y-%m").strftime("%b %Y"), H_FILL, H_FONT, "center")
    styled_cell(ws, row, N_COLS, "ИТОГО", H_FILL, H_FONT, "center")

    row += 1
    ws.merge_cells(f"A{row}:{get_column_letter(N_COLS)}{row}")
    ws.cell(row=row, column=1, value="💚 ПОСТУПЛЕНИЯ").fill = IN_FILL
    ws.cell(row=row, column=1).font = BOLD

    for cat in income_cats:
        row += 1
        styled_cell(ws, row, 1, cat, IN_FILL)
        total = 0
        for i, m in enumerate(months):
            v = max(monthly[m].get(cat, 0), 0)
            styled_cell(ws, row, i+2, v, num_fmt='# ##0.00 ₽')
            total += v
        styled_cell(ws, row, N_COLS, total, font=BOLD, num_fmt='# ##0.00 ₽')

    row += 1
    styled_cell(ws, row, 1, "Итого поступления", TOT_FILL, BOLD)
    total_income = 0
    for i, m in enumerate(months):
        v = sum(a for a in monthly[m].values() if a > 0)
        styled_cell(ws, row, i+2, v, TOT_FILL, BOLD, num_fmt='# ##0.00 ₽')
        total_income += v
    styled_cell(ws, row, N_COLS, total_income, TOT_FILL, BOLD, num_fmt='# ##0.00 ₽')

    row += 2
    ws.merge_cells(f"A{row}:{get_column_letter(N_COLS)}{row}")
    ws.cell(row=row, column=1, value="🔴 РАСХОДЫ").fill = EX_FILL
    ws.cell(row=row, column=1).font = BOLD

    for cat in expense_cats:
        row += 1
        styled_cell(ws, row, 1, cat, EX_FILL)
        total = 0
        for i, m in enumerate(months):
            v = abs(monthly[m].get(cat, 0)) if monthly[m].get(cat, 0) < 0 else 0
            styled_cell(ws, row, i+2, v, num_fmt='# ##0.00 ₽')
            total += v
        styled_cell(ws, row, N_COLS, total, font=BOLD, num_fmt='# ##0.00 ₽')

    row += 1
    styled_cell(ws, row, 1, "Итого расходы", TOT_FILL, BOLD)
    total_expense = 0
    for i, m in enumerate(months):
        v = abs(sum(a for a in monthly[m].values() if a < 0))
        styled_cell(ws, row, i+2, v, TOT_FILL, BOLD, num_fmt='# ##0.00 ₽')
        total_expense += v
    styled_cell(ws, row, N_COLS, total_expense, TOT_FILL, BOLD, num_fmt='# ##0.00 ₽')

    row += 2
    styled_cell(ws, row, 1, "💰 ЧИСТЫЙ ДЕНЕЖНЫЙ ПОТОК", NDP_FILL, BOLD)
    for i, m in enumerate(months):
        ndp = sum(monthly[m].values())
        c = styled_cell(ws, row, i+2, ndp, NDP_FILL, num_fmt='# ##0.00 ₽')
        c.font = Font(bold=True, color="C00000" if ndp < 0 else "000000")
    total_ndp = total_income - total_expense
    c = styled_cell(ws, row, N_COLS, total_ndp, NDP_FILL, num_fmt='# ##0.00 ₽')
    c.font = Font(bold=True, color="C00000" if total_ndp < 0 else "000000")

    ws.column_dimensions["A"].width = 32
    for i in range(len(months)+1):
        ws.column_dimensions[get_column_letter(i+2)].width = 16

def build_transactions_sheet(ws, transactions):
    H_FILL = PatternFill("solid", fgColor="1F4E79")
    for col, h in enumerate(["Дата", "Описание", "Категория", "Сумма", "Тип"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = H_FILL
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
    for row, t in enumerate(sorted(transactions, key=lambda x: x["date"]), 2):
        ws.cell(row=row, column=1, value=t["date"].strftime("%d.%m.%Y"))
        ws.cell(row=row, column=2, value=t["description"][:100])
        ws.cell(row=row, column=3, value=t.get("category", ""))
        c = ws.cell(row=row, column=4, value=t["amount"])
        c.number_format = '# ##0.00 ₽'
        c.font = Font(color="C00000" if t["amount"] < 0 else "375623")
        ws.cell(row=row, column=5, value="Поступление" if t["amount"] > 0 else "Расход")
    for col, w in zip("ABCDE", [12, 50, 30, 16, 14]):
        ws.column_dimensions[col].width = w

def build_summary_text(transactions, bank_name):
    total_income = sum(t["amount"] for t in transactions if t["amount"] > 0)
    total_expense = abs(sum(t["amount"] for t in transactions if t["amount"] < 0))
    ndp = total_income - total_expense
    exp_by_cat = defaultdict(float)
    for t in transactions:
        if t["amount"] < 0:
            exp_by_cat[t.get("category", "Прочее")] += abs(t["amount"])
    top = sorted(exp_by_cat.items(), key=lambda x: x[1], reverse=True)[:3]
    d_from = min(t["date"] for t in transactions).strftime("%d.%m.%Y")
    d_to = max(t["date"] for t in transactions).strftime("%d.%m.%Y")
    lines = [
        f"📊 <b>ДДС-отчёт | {bank_name}</b>",
        f"📅 Период: {d_from} — {d_to}",
        f"📋 Транзакций: {len(transactions)}",
        "",
        f"💚 Поступления: <b>{total_income:,.2f} ₽</b>",
        f"🔴 Расходы: <b>{total_expense:,.2f} ₽</b>",
        f"{'💰' if ndp >= 0 else '⚠️'} Чистый поток: <b>{ndp:+,.2f} ₽</b>",
        "", "📌 <b>Топ расходов:</b>",
    ]
    for cat, amt in top:
        pct = amt / total_expense * 100 if total_expense else 0
        lines.append(f"  • {cat}: {amt:,.0f} ₽ ({pct:.0f}%)")
    return "\n".join(lines)
